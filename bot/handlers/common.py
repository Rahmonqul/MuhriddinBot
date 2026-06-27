import logging

from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery
from asgiref.sync import sync_to_async

from bot.keyboards.reply import main_keyboard, admin_keyboard

logger = logging.getLogger(__name__)
common_router = Router()


def _is_admin(telegram_id: int) -> bool:
    from django.conf import settings
    return telegram_id in settings.ADMIN_TELEGRAM_IDS


# ─── Test xabar yuborish ─────────────────────────────────────────────────────

@common_router.message(Command('test_notify'))
async def test_notify(message: Message):
    """Admin uchun: /test_notify — xabarnoma ishlayotganini tekshirish"""
    if not _is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    try:
        await message.bot.send_message(
            message.from_user.id,
            "✅ *Test xabarnomasi ishlaydi!*\n\n"
            "Endi ro'yxatdan o'tganda xabarlar keladi.",
            parse_mode='Markdown'
        )
        logger.info("Test xabarnomasi muvaffaqiyatli yuborildi.")
        await message.answer("✅ Xabarnoma tizimi ishlayapti!")
    except Exception as e:
        logger.error("Test xabarnomasi yuborishda xatolik: %s", e)
        await message.answer(f"❌ Xatolik: {e}")


# ─── Admin buyruqlari ────────────────────────────────────────────────────────

@common_router.message(F.text == "👥 O'quvchilar holati")
async def admin_students_status(message: Message):
    """O'quvchilar soni (guruh bo'yicha) + holat (rang) — bitta xabar."""
    if not _is_admin(message.from_user.id):
        return

    from academy.models import Student, Group, get_student_status
    from django.db.models import Count, Q as DjQ

    # Guruh bo'yicha sonlar
    groups = await sync_to_async(
        lambda: list(
            Group.objects.filter(is_active=True)
            .annotate(cnt=Count('students', filter=DjQ(students__is_active=True)))
            .order_by('name')
        )
    )()
    total = await sync_to_async(Student.objects.filter(is_active=True).count)()
    no_group = await sync_to_async(
        Student.objects.filter(is_active=True, group__isnull=True).count
    )()

    text = "👥 *O'quvchilar holati:*\n\n"
    for g in groups:
        text += f"🏫 *{g.name}:* {g.cnt} ta\n"
    text += f"\nJami faol: *{total}* ta"
    if no_group:
        text += f" | Guruhsiz: *{no_group}* ta"
    text += f"\n\n{'─' * 22}\n🎨 *Holat bo'yicha:*\n\n"

    # Status hisoblash
    students = await sync_to_async(
        lambda: list(
            Student.objects.filter(is_active=True)
            .select_related('group')
            .prefetch_related('attendances', 'payments', 'test_results__test')
        )
    )()

    counts = {'green': 0, 'yellow': 0, 'red': 0, 'black': 0}
    critical = []
    for s in students:
        key, emoji, label = await sync_to_async(get_student_status)(s)
        counts[key] += 1
        if key in ('red', 'black'):
            group_name = s.group.name if s.group else '—'
            critical.append(f"  {emoji} {s.full_name} ({group_name})")

    text += (
        f"🟢 Yaxshi:     *{counts['green']}* ta\n"
        f"🟡 O'rta:      *{counts['yellow']}* ta\n"
        f"🔴 Yomon:      *{counts['red']}* ta\n"
        f"⚫ Juda yomon: *{counts['black']}* ta\n"
    )
    if critical:
        text += "\n⚠️ *Diqqat talab qiladiganlar:*\n"
        text += "\n".join(critical[:12])
        if len(critical) > 12:
            text += f"\n  ... va yana {len(critical) - 12} ta"

    await message.answer(text, parse_mode='Markdown', reply_markup=admin_keyboard())

    # Inline keyboard — Excel yuklab olish uchun guruh tugmalari
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    excel_buttons = [
        [InlineKeyboardButton(
            text=f"📥 {g.name} — Excel ({g.cnt} ta)",
            callback_data=f"grp_excel_{g.pk}"
        )]
        for g in groups if g.cnt > 0
    ]
    if excel_buttons:
        await message.answer(
            "📊 Guruh bo'yicha Excel yuklash:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=excel_buttons)
        )


@common_router.message(F.text == "📊 Guruhlar & Statistika")
async def admin_groups_stats(message: Message):
    """Guruhlar ro'yxati + umumiy statistika — bitta xabar."""
    if not _is_admin(message.from_user.id):
        return

    from academy.models import Group, Student, Payment
    from django.db.models import Count, Q, Sum

    groups = await sync_to_async(
        lambda: list(Group.objects.filter(is_active=True).select_related('teacher').order_by('name'))
    )()

    text = "📋 *Guruhlar:*\n\n"
    total_students = 0
    for g in groups:
        cnt = await sync_to_async(Student.objects.filter(group=g, is_active=True).count)()
        total_students += cnt
        teacher = g.teacher.full_name if g.teacher else "Belgilanmagan"
        fee = f"{g.monthly_fee:,.0f} so'm" if g.monthly_fee else "—"
        text += (
            f"🏫 *{g.name}* — {cnt} ta o'quvchi\n"
            f"   👨‍🏫 {teacher} | 💰 {fee}\n\n"
        )
    if not groups:
        text += "_Guruhlar yo'q_\n\n"

    # Umumiy statistika
    no_group = await sync_to_async(
        Student.objects.filter(is_active=True, group__isnull=True).count
    )()
    debtors = await sync_to_async(
        lambda: Student.objects.filter(is_active=True, payments__status='overdue').distinct().count()
    )()
    pay_stats = await sync_to_async(
        lambda: Payment.objects.aggregate(
            paid=Count('id', filter=Q(status='paid')),
            pending=Count('id', filter=Q(status='pending')),
            overdue=Count('id', filter=Q(status='overdue')),
            paid_sum=Sum('paid_amount', filter=Q(status='paid')),
        )
    )()

    text += (
        f"{'─' * 22}\n"
        f"📊 *Umumiy statistika:*\n\n"
        f"👥 Faol o'quvchilar: *{total_students}* ta\n"
    )
    if no_group:
        text += f"📌 Guruhsiz: *{no_group}* ta\n"
    text += (
        f"🔴 Qarzdorlar: *{debtors}* ta\n\n"
        f"💳 *To'lovlar:*\n"
        f"   ✅ To'langan: *{pay_stats['paid']}* ta\n"
        f"   ⏳ Kutilmoqda: *{pay_stats['pending']}* ta\n"
        f"   🔴 Muddati o'tdi: *{pay_stats['overdue']}* ta\n"
        f"   💵 Yig'ilgan: *{(pay_stats['paid_sum'] or 0):,.0f} so'm*"
    )

    await message.answer(text, parse_mode='Markdown', reply_markup=admin_keyboard())


async def get_student(telegram_id: int):
    from academy.models import Student
    from django.db.models import Q
    return await sync_to_async(
        lambda: Student.objects.filter(
            Q(telegram_id=telegram_id) | Q(parent_telegram_id=telegram_id)
        ).select_related('group').first()
    )()


@common_router.message(F.text == "📅 Dars jadvali")
async def show_schedule(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return
    if not student.group:
        await message.answer(
            "📋 Sizga hali guruh belgilanmagan.\n"
            "Administrator tez orada biriktiriladi. ⏳"
        )
        return

    from academy.models import Schedule
    schedules = await sync_to_async(
        lambda: list(
            Schedule.objects.filter(group=student.group).order_by('start_time')
        )
    )()

    if not schedules:
        await message.answer("Guruhingiz uchun dars jadvali hali belgilanmagan.")
        return

    DAYS = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba']
    text = f"📅 *{student.group.name}* guruhi dars jadvali:\n\n"
    for s in schedules:
        room = f"  🚪 {s.room}" if s.room else ""
        days_str = ', '.join(DAYS[d] for d in sorted(s.days_of_week))
        text += (
            f"📌 *{days_str}*\n"
            f"   🕐 {s.start_time.strftime('%H:%M')} — {s.end_time.strftime('%H:%M')}{room}\n\n"
        )

    await message.answer(text, parse_mode='Markdown')


@common_router.message(F.text.in_({"💰 To'lovlar", "💳 To'lov jadvali"}))
async def show_payments(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return

    from academy.models import Payment
    payments = await sync_to_async(
        lambda: list(Payment.objects.filter(student=student).order_by('-month')[:6])
    )()

    if not payments:
        await message.answer("📋 To'lov ma'lumotlari hali yo'q.")
        return

    months_uz = [
        '', 'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun',
        'Iyul', 'Avgust', 'Sentyabr', 'Oktyabr', 'Noyabr', 'Dekabr'
    ]
    emoji_map = {'paid': '✅', 'pending': '⏳', 'overdue': '🔴', 'partial': '🟡'}

    text = "💰 *To'lovlar tarixi:*\n\n"
    for p in payments:
        em = emoji_map.get(p.status, '❓')
        month_name = f"{months_uz[p.month.month]} {p.month.year}"
        text += f"{em} *{month_name}*\n"
        text += f"   💵 {p.amount:,.0f} so'm\n"
        text += f"   📅 Muddat: {p.due_date.strftime('%d.%m.%Y')}\n"
        if p.paid_date:
            text += f"   ✅ To'langan: {p.paid_date.strftime('%d.%m.%Y')}\n"
        text += "\n"

    await message.answer(text, parse_mode='Markdown')


@common_router.message(F.text == "📊 Davomatim")
async def show_attendance(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return

    from academy.models import Attendance
    from django.db.models import Count, Q

    stats = await sync_to_async(
        lambda: Attendance.objects.filter(student=student).aggregate(
            total=Count('id'),
            present=Count('id', filter=Q(status='present')),
            late=Count('id', filter=Q(status='late')),
            absent=Count('id', filter=Q(status='absent')),
        )
    )()

    total = stats['total']
    if total == 0:
        await message.answer("📊 Davomat ma'lumotlari hali yo'q.")
        return

    attended = stats['present'] + stats['late']
    pct = int(attended / total * 100)

    progress = '🟩' * (pct // 10) + '⬜' * (10 - pct // 10)

    recent = await sync_to_async(
        lambda: list(
            Attendance.objects.filter(student=student)
            .select_related('lesson__schedule__group')
            .order_by('-lesson__date')[:5]
        )
    )()

    text = (
        f"📊 *Davomat hisoboti:*\n\n"
        f"{progress} {pct}%\n\n"
        f"✅ Keldi: {stats['present']} ta\n"
        f"⏰ Kech qoldi: {stats['late']} ta\n"
        f"❌ Kelmadi: {stats['absent']} ta\n"
        f"📝 Jami: {total} ta dars\n\n"
    )

    if recent:
        text += "*So'nggi darslar:*\n"
        em_map = {'present': '✅', 'late': '⏰', 'absent': '❌'}
        for a in recent:
            text += f"{em_map[a.status]} {a.lesson.date.strftime('%d.%m.%Y')}\n"

    await message.answer(text, parse_mode='Markdown')


@common_router.message(F.text == "👤 Profilim")
async def show_profile(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return

    group_name = student.group.name if student.group else "Belgilanmagan"

    overdue_count = await sync_to_async(
        student.payments.filter(status='overdue').count
    )()

    debt_line = (
        f"⚠️ Qarzdorlik: {overdue_count} ta to'lov\n"
        if overdue_count else "✅ Qarzdorlik yo'q\n"
    )

    # Umumiy kechikish daqiqalari
    from django.db.models import Sum, Count
    late_stats = await sync_to_async(
        lambda: student.attendances.filter(status='late').aggregate(
            total_min=Sum('late_minutes'),
            count=Count('id'),
        )
    )()
    total_late = late_stats['total_min'] or 0
    late_count = late_stats['count'] or 0

    if total_late > 0:
        hours = total_late // 60
        mins  = total_late % 60
        time_str = f"{hours}s {mins}d" if hours else f"{mins} daqiqa"
        late_line = f"⏰ Kechikish: {late_count} marta, jami *{time_str}*\n"
    else:
        late_line = "⏰ Kechikish: yo'q ✓\n"

    status_text = await _student_status_text(student)

    text = (
        f"👤 *Mening profilim:*\n\n"
        f"📝 Ism: {student.full_name}\n"
        f"📞 Telefon: {student.phone}\n"
        f"🏫 Guruh: {group_name}\n"
        f"{debt_line}"
        f"{late_line}"
        f"📅 Ro'yxat: {student.registered_at.strftime('%d.%m.%Y')}\n\n"
        f"{status_text}"
    )

    await message.answer(text, parse_mode='Markdown', reply_markup=main_keyboard())


# ─── Nazorat ishlari ─────────────────────────────────────────────────────────

@common_router.message(F.text == "📝 Nazorat ishlari")
async def show_test_results(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return

    if not student.group_id:
        await message.answer("📋 Sizga hali guruh belgilanmagan. ⏳")
        return

    from academy.models import TestResult, Certificate
    results = await sync_to_async(
        lambda: list(
            TestResult.objects
            .filter(student=student)
            .select_related('test')
            .order_by('-test__date')[:10]
        )
    )()

    if not results:
        await message.answer(
            "📝 Hali nazorat ishi natijalari kiritilmagan.\n"
            "Natijalar admin tomonidan kiritilgach ko'rinadi."
        )
        return

    result_ids = [r.pk for r in results]
    cert_map = await sync_to_async(
        lambda: dict(
            Certificate.objects
            .filter(test_result_id__in=result_ids)
            .values_list('test_result_id', 'grade')
        )
    )()

    grade_emoji = {'A+': '🌟', 'A': '⭐', 'B+': '💫', 'B': '✨', 'C+': '🎖️', 'C': '🎗️'}

    text = "📝 *Nazorat ishlari natijalari:*\n\n"
    for r in results:
        pct = int(r.score / r.test.max_score * 100)
        if pct >= 70:
            em = "🟢"
        elif pct >= 50:
            em = "🟡"
        else:
            em = "🔴"

        if r.rank:
            grade = cert_map.get(r.pk)
            if grade:
                em_g = grade_emoji.get(grade, '🏅')
                rank_info = f"{r.rank}-o'rin | {em_g} *{grade}* sertifikat"
            else:
                rank_info = f"{r.rank}-o'rin"
        elif r.test.min_score > 0 and r.score < r.test.min_score:
            rank_info = f"❌ Min. ball ({r.test.min_score}) yetmadi"
        else:
            rank_info = "—"

        text += (
            f"{em} *{r.test.title}*\n"
            f"   📅 {r.test.date.strftime('%d.%m.%Y')}\n"
            f"   🎯 Ball: *{r.score} / {r.test.max_score}* ({pct}%)\n"
            f"   🏆 {rank_info}\n\n"
        )

    await message.answer(text, parse_mode='Markdown')


# ─── O'quvchi holati (rang/status) ───────────────────────────────────────────

async def _student_status_text(student) -> str:
    from academy.models import get_student_status
    key, emoji, label = await sync_to_async(get_student_status)(student)

    descs = {
        'green':  "Davomat, nazorat ishlari va to'lovlar yaxshi. Shunday davom eting! 💪",
        'yellow': "Bir yoki bir nechta ko'rsatkich o'rtacha. E'tibor bering! ⚠️",
        'red':    "Ko'rsatkichlar yomon. O'qituvchi bilan gaplashing. ❗",
        'black':  "Jiddiy muammolar bor: qarzdorlik, kam davomat yoki past ball. 🆘",
    }
    return f"{emoji} *Sizning holatingiz: {label}*\n\n{descs[key]}"


# ─── Admin: to'lov hisoboti (oylik, guruh bo'yicha) ─────────────────────────

MONTHS_UZ = ['', 'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun',
             'Iyul', 'Avgust', 'Sentyabr', 'Oktyabr', 'Noyabr', 'Dekabr']


@common_router.message(F.text == "💰 Barcha to'lovlar")
async def admin_all_payments(message: Message):
    if not _is_admin(message.from_user.id):
        return
    await _send_payment_report(message, months_back=0)


async def _send_payment_report(message, months_back: int = 0):
    from academy.models import Group, Payment
    from django.db.models import Count, Q, Sum
    from datetime import date, timedelta

    today = date.today()
    # months_back=0 → joriy oy, 1 → o'tgan oy
    year  = today.year
    month = today.month - months_back
    if month <= 0:
        month += 12
        year  -= 1
    month_start = date(year, month, 1)

    header_month = f"{MONTHS_UZ[month]} {year}"

    groups = await sync_to_async(
        lambda: list(Group.objects.filter(is_active=True).order_by('name'))
    )()

    total_paid_sum    = 0
    total_pending_sum = 0
    total_overdue_sum = 0
    total_upcoming    = 0
    upcoming_cutoff   = today + timedelta(days=3)

    text = f"💰 *{header_month} — To'lov hisoboti*\n\n"

    for group in groups:
        stats = await sync_to_async(
            lambda g=group: Payment.objects.filter(
                student__group=g, month=month_start
            ).aggregate(
                total=Count('id'),
                paid_cnt=Count('id', filter=Q(status='paid')),
                pending_cnt=Count('id', filter=Q(status='pending')),
                overdue_cnt=Count('id', filter=Q(status='overdue')),
                paid_sum=Sum('paid_amount', filter=Q(status='paid')),
                pending_sum=Sum('amount', filter=Q(status='pending')),
                overdue_sum=Sum('amount', filter=Q(status='overdue')),
                upcoming_cnt=Count('id', filter=Q(
                    status='pending',
                    due_date__range=[today, upcoming_cutoff],
                )),
            )
        )()

        if stats['total'] == 0:
            continue

        ps  = stats['paid_sum']    or 0
        pns = stats['pending_sum'] or 0
        os_ = stats['overdue_sum'] or 0

        total_paid_sum    += ps
        total_pending_sum += pns
        total_overdue_sum += os_
        total_upcoming    += stats['upcoming_cnt']

        group_line = f"🏫 *{group.name}*\n"
        if stats['paid_cnt']:
            group_line += f"   ✅ To'landi: {stats['paid_cnt']} ta — {ps:,.0f} so'm\n"
        if stats['pending_cnt']:
            group_line += f"   ⏳ Kutmoqda: {stats['pending_cnt']} ta — {pns:,.0f} so'm\n"
        if stats['overdue_cnt']:
            group_line += f"   🔴 Muddati o'tdi: {stats['overdue_cnt']} ta — {os_:,.0f} so'm\n"
        if stats['upcoming_cnt']:
            group_line += f"   ⏰ 3 kun ichida: {stats['upcoming_cnt']} ta\n"
        text += group_line + "\n"

    # Jami
    total_sum = total_paid_sum + total_pending_sum + total_overdue_sum
    text += (
        f"{'━' * 20}\n"
        f"📊 *Jami {header_month}:*\n"
        f"   💵 Yig'ilgan:      *{total_paid_sum:,.0f} so'm*\n"
        f"   ⏳ Kutilayotgan:   *{total_pending_sum:,.0f} so'm*\n"
        f"   🔴 Qarzdorlik:     *{total_overdue_sum:,.0f} so'm*\n"
        f"   🎯 Umumiy kutilgan: *{total_sum:,.0f} so'm*\n"
    )
    if total_upcoming:
        text += f"   ⏰ Yaqin 3 kun:    *{total_upcoming} ta to'lov*\n"

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    nav_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="◀️ O'tgan oy",
            callback_data=f"pay_report_{months_back + 1}"
        ),
        InlineKeyboardButton(
            text="▶️ Joriy oy",
            callback_data=f"pay_report_{max(months_back - 1, 0)}"
        ),
    ]])

    await message.answer(text, parse_mode='Markdown', reply_markup=nav_kb)


@common_router.callback_query(F.data.startswith("pay_report_"))
async def cb_payment_report(callback: CallbackQuery):
    if not _is_admin(callback.from_user.id):
        await callback.answer()
        return
    months_back = int(callback.data.split("_")[-1])
    await callback.answer()
    try:
        await callback.message.delete()
    except Exception:
        pass
    await _send_payment_report(callback.message, months_back=months_back)


# ─── Admin: guruh Excel eksport ──────────────────────────────────────────────

async def _generate_group_excel(group_id: int) -> bytes:
    import io
    from datetime import timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from academy.models import Student, get_student_status
    from django.db.models import Sum, Count, Q

    today = __import__('datetime').date.today()
    month_ago = today - timedelta(days=30)

    students = await sync_to_async(
        lambda: list(
            Student.objects.filter(group_id=group_id, is_active=True)
            .select_related('group')
            .order_by('full_name')
        )
    )()

    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    headers = [
        '№', "To'liq ism", 'Telefon', 'Telegram', "Tug'ilgan sana",
        'Guruh', "Ro'yxat sanasi", 'Holat',
        'Davomat %', 'Kechikish (min)',
        "Qarzdorlik (so'm)", "Kutilayotgan to'lov (so'm)",
        'Sertifikatlar',
    ]

    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    status_fills = {
        'green':  PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'yellow': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'red':    PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'black':  PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    }
    status_labels = {
        'green': 'Yaxshi', 'yellow': "O'rta", 'red': 'Yomon', 'black': 'Juda yomon',
    }

    for i, student in enumerate(students, 2):
        # Davomat
        att = await sync_to_async(
            lambda s=student: s.attendances.filter(lesson__date__gte=month_ago).aggregate(
                total=Count('id'),
                ok=Count('id', filter=Q(status__in=['present', 'late'])),
                late_sum=Sum('late_minutes'),
            )
        )()
        att_pct = int(att['ok'] / att['total'] * 100) if att['total'] else 0

        # To'lovlar
        pay = await sync_to_async(
            lambda s=student: s.payments.aggregate(
                debt=Sum('amount', filter=Q(status='overdue')),
                pending=Sum('amount', filter=Q(status='pending')),
            )
        )()

        # Sertifikatlar
        certs = await sync_to_async(
            lambda s=student: ', '.join(
                s.certificates.order_by('-issued_date').values_list('grade', flat=True)
            ) or '—'
        )()

        # Status
        key, _, _ = await sync_to_async(get_student_status)(student)

        tg = f"@{student.telegram_username}" if student.telegram_username else (
            str(student.telegram_id) if student.telegram_id else '—'
        )
        bdate = student.birth_date.strftime('%d.%m.%Y') if student.birth_date else '—'
        group = student.group.name if student.group else '—'

        row = [
            i - 1,
            student.full_name,
            student.phone,
            tg,
            bdate,
            group,
            student.registered_at.strftime('%d.%m.%Y'),
            status_labels.get(key, key),
            f"{att_pct}%",
            att['late_sum'] or 0,
            int(pay['debt'] or 0),
            int(pay['pending'] or 0),
            certs,
        ]

        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = left if col == 2 else center
            if col == 8:
                c.fill = status_fills.get(key, PatternFill())
        ws.row_dimensions[i].height = 18

    col_widths = [5, 28, 16, 20, 14, 14, 14, 14, 10, 14, 18, 22, 22]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@common_router.callback_query(F.data.startswith("grp_excel_"))
async def cb_group_excel(callback: CallbackQuery):
    if not _is_admin(callback.from_user.id):
        await callback.answer()
        return

    group_id = int(callback.data.split("_")[-1])
    await callback.answer("⏳ Excel tayyorlanmoqda...")

    from academy.models import Group
    from aiogram.types import BufferedInputFile

    group = await sync_to_async(Group.objects.select_related('teacher').get)(pk=group_id)

    try:
        excel_bytes = await _generate_group_excel(group_id)
    except Exception as e:
        logger.error("Excel generation error: %s", e)
        await callback.message.answer(f"❌ Excel yaratishda xatolik: {e}")
        return

    today_str = __import__('datetime').date.today().strftime('%d.%m.%Y')
    filename = f"{group.name}_oquvchilar_{today_str}.xlsx"

    await callback.message.answer_document(
        BufferedInputFile(excel_bytes, filename=filename),
        caption=(
            f"📊 *{group.name}* guruhi — o'quvchilar ro'yxati\n"
            f"📅 Sana: {today_str}"
        ),
        parse_mode='Markdown',
    )


# ─── Admin: o'quvchilar holati (rang) ────────────────────────────────────────

@common_router.message(F.text == "🎨 O'quvchilar holati")
async def admin_student_statuses(message: Message):
    if not _is_admin(message.from_user.id):
        return

    from academy.models import Student, get_student_status

    students = await sync_to_async(
        lambda: list(
            Student.objects.filter(is_active=True)
            .select_related('group')
            .prefetch_related('attendances', 'payments', 'test_results__test')
        )
    )()

    counts = {'green': 0, 'yellow': 0, 'red': 0, 'black': 0}
    critical = []

    for s in students:
        key, emoji, label = await sync_to_async(get_student_status)(s)
        counts[key] += 1
        if key in ('red', 'black'):
            critical.append(f"  {emoji} {s.full_name} ({s.group.name if s.group else '—'})")

    text = (
        "🎨 *O'quvchilar holati:*\n\n"
        f"🟢 Yaxshi: *{counts['green']}* ta\n"
        f"🟡 O'rta: *{counts['yellow']}* ta\n"
        f"🔴 Yomon: *{counts['red']}* ta\n"
        f"⚫ Juda yomon: *{counts['black']}* ta\n"
    )
    if critical:
        text += "\n*Diqqat talab qiladigan o'quvchilar:*\n"
        text += "\n".join(critical[:15])
        if len(critical) > 15:
            text += f"\n... va yana {len(critical) - 15} ta"

    await message.answer(text, parse_mode='Markdown', reply_markup=admin_keyboard())


# ─── Yutuqlarim (Sertifikatlar) ───────────────────────────────────────────────

@common_router.message(F.text == "🏆 Yutuqlarim")
async def show_achievements(message: Message):
    student = await get_student(message.from_user.id)
    if not student:
        await message.answer("Avval ro'yxatdan o'ting! /start")
        return

    from academy.models import Certificate
    certs = await sync_to_async(
        lambda: list(
            Certificate.objects
            .filter(student=student)
            .select_related('test_result__test')
            .order_by('-issued_date', 'rank')
        )
    )()

    if not certs:
        await message.answer(
            "🏆 *Yutuqlarim*\n\n"
            "Hali sertifikat yo'q.\n"
            "Nazorat ishlarida yaxshi natija ko'rsatganda sertifikat qo'shiladi! 💪",
            parse_mode='Markdown'
        )
        return

    medals = {1: '🥇', 2: '🥈', 3: '🥉'}
    grade_emoji = {'A+': '🌟', 'A': '⭐', 'B+': '💫', 'B': '✨', 'C+': '🎖️', 'C': '🎗️'}

    text = f"🏆 *Mening sertifikatlarim:* {len(certs)} ta\n\n"
    for c in certs:
        em = grade_emoji.get(c.grade, '🏅')
        medal = medals.get(c.rank, '')
        test = c.test_result.test
        pct = int(c.test_result.score / test.max_score * 100)
        text += (
            f"{em} *{c.grade} sertifikat* — {test.title}\n"
            f"   {medal} O'rin: *{c.rank}-o'rin*\n"
            f"   🎯 Ball: *{c.test_result.score}/{test.max_score}* ({pct}%)\n"
            f"   📅 Sana: *{c.issued_date.strftime('%d.%m.%Y')}*\n\n"
        )

    await message.answer(text, parse_mode='Markdown')


@common_router.message(Command('help'))
async def cmd_help(message: Message):
    await message.answer(
        "ℹ️ *MathAcademy Bot — Yordam*\n\n"
        "📅 *Dars jadvali* — haftalik dars vaqtlarini ko'ring\n"
        "💰 *To'lovlar* — to'lov tarixingizni ko'ring\n"
        "📊 *Davomatim* — davomat statistikangiz\n"
        "👤 *Profilim* — shaxsiy ma'lumotlaringiz\n\n"
        "Muammo bo'lsa, administrator bilan bog'laning.",
        parse_mode='Markdown'
    )
